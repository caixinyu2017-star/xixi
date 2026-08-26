# -*- coding: utf-8 -*-
"""Write the simulated dataset as an Excel workbook.

Sheet 1 says what the file is, in the largest type on the page, because a
spreadsheet gets forwarded and the label has to travel with it. Sheet 2 is
the codebook. Sheet 3 is the data, in the column order Wenjuanxing exports.
"""
import json
import os
import sys

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.abspath(os.path.join(HERE, ".."))
for p in ("design", "sim"):
    sys.path.insert(0, os.path.join(ROOT, p))
import items as I                                             # noqa: E402
import simulate as SIM                                        # noqa: E402

OUT = os.path.join(ROOT, "模拟数据_SIMULATED_N400.xlsx")

RED = Font(name="微软雅黑", size=11, bold=True, color="9C0000")
H1 = Font(name="微软雅黑", size=15, bold=True)
HDR = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
BODY = Font(name="微软雅黑", size=10)
FILL = PatternFill("solid", fgColor="44546A")
WARN = PatternFill("solid", fgColor="FDE9E9")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")


def build():
    data, _, meta = SIM.simulate()
    wb = Workbook()

    # ------------------------------------------------------- sheet 1
    ws = wb.active
    ws.title = "务必先读"
    ws.column_dimensions["A"].width = 104
    lines = [
        ("本文件是模拟数据，不是真实调查数据。", H1, WARN),
        ("", BODY, None),
        ("文件中的 400 行作答由 sim/simulate.py 按研究假设的结构方程生成，"
         "没有任何一位真实被试参与。", RED, WARN),
        ("", BODY, None),
        ("它的用途只有三个：", BODY, None),
        ("　1. 让分析脚本在拿到真实数据之前就能跑通并定稿；", BODY, None),
        ("　2. 让论文的表格结构、变量命名和呈现方式在开工前就确定下来；",
         BODY, None),
        ("　3. 做检验力分析，判断计划的样本量能否检出所假设的效应。", BODY,
         None),
        ("", BODY, None),
        ("它不能做的事：", BODY, None),
        ("　·　不能用来支持任何实证结论；", RED, None),
        ("　·　不能写进论文的结果部分；", RED, None),
        ("　·　不能提交给期刊、审稿人或数据仓储作为研究数据。", RED, None),
        ("", BODY, None),
        ("数据是按假设生成的，所以它当然“支持”假设——这既是它的用途，"
         "也是它的界限。真实数据回收后，请用同一套脚本重跑，"
         "并以真实结果为准，无论方向是否与此处一致。", BODY, None),
        ("", BODY, None),
        ("生成参数", H1, None),
        ("随机种子：%d" % meta["seed"], BODY, None),
        ("模拟发放：%d 份；通过筛选后保留：%d 份"
         % (meta["n_collected"], meta["n_valid"]), BODY, None),
        ("剔除：%s" % "；".join("%s %d 份" % (k, v)
                                for k, v in meta["excluded"].items()),
         BODY, None),
        ("生成用结构系数：%s"
         % json.dumps(meta["structural"], ensure_ascii=False), BODY, None),
        ("", BODY, None),
        ("生成脚本：sim/simulate.py　　分析脚本：analysis/run.py", BODY, None),
    ]
    for i, (t, f, fill) in enumerate(lines, 1):
        c = ws.cell(row=i, column=1, value=t)
        c.font = f
        c.alignment = Alignment(wrap_text=True, vertical="center")
        if fill:
            c.fill = fill
        ws.row_dimensions[i].height = 34 if f is H1 else 20
    ws.sheet_view.showGridLines = False

    # ------------------------------------------------------- sheet 2
    ws = wb.create_sheet("变量说明")
    head = ["变量名", "所属部分", "题目 / 含义", "取值", "反向计分"]
    for j, h in enumerate(head, 1):
        c = ws.cell(row=1, column=j, value=h)
        c.font, c.fill, c.border = HDR, FILL, BOX
    r = 2

    def add(name, part, text, vals, rev=""):
        nonlocal r
        for j, v in enumerate([name, part, text, vals, rev], 1):
            c = ws.cell(row=r, column=j, value=v)
            c.font, c.border, c.alignment = BODY, BOX, WRAP
        r += 1

    add("ID", "编号", "个案序号", "1–%d" % meta["n_valid"])
    lab = {c: (n, o) for c, n, o in I.DEMO}
    for code in I.CONTROLS:
        n_, o_ = lab[code]
        add(code, "第一部分  基本情况", n_, "；".join(o_))
    for con in I.CONSTRUCTS:
        txt, k = I.SCALES[con.scale]
        for code, t, rv in con.items:
            add(code, "第二部分  %s" % con.cn, t, txt, "是" if rv else "")
    for code, t, sc, ok, _ in I.ATTENTION:
        add(code, "注意力检测", t, I.SCALES[sc][0], "正确答案 %d" % ok)
    add("duration_s", "系统记录", "作答时长（秒）", "整数")
    for j, w in enumerate([13, 20, 46, 40, 14], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A2"

    # ------------------------------------------------------- sheet 3
    ws = wb.create_sheet("数据_SIMULATED")
    cols = SIM.ORDER
    for j, name in enumerate(cols, 1):
        c = ws.cell(row=1, column=j, value=name)
        c.font, c.fill, c.border = HDR, FILL, BOX
        ws.column_dimensions[get_column_letter(j)].width = \
            max(7, min(13, len(name) + 3))
    arr = {k: np.asarray(data[k]) for k in cols}
    for i in range(meta["n_valid"]):
        for j, name in enumerate(cols, 1):
            ws.cell(row=i + 2, column=j, value=int(arr[name][i])).font = BODY
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(cols)),
                                      meta["n_valid"] + 1)

    wb.save(OUT)
    print("saved:", OUT)
    print("  %d 行 × %d 列" % (meta["n_valid"], len(cols)))


if __name__ == "__main__":
    build()
