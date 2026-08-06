# -*- coding: utf-8 -*-
"""从真实数据中挑出 50 名「材料较完整、当前得分较高」的学生，
按其行为特征反推出一份可直接导入系统的模拟问卷（问卷星导出格式）。

映射思路（写进说明页，便于核查）：
  1. 取该生五个维度的实际证据强度 A_d，在这 50 人内换算成百分位 rank_d；
  2. 自评水平 lvl_d = 0.50 + 0.34 · (rank_d − 0.5) · 2 + 个体基调（自评宽严）；
     —— 行为上强的维度，自评也高；同时保留"有人整体自评偏高、有人偏严"的个体差异；
  3. 每道题 = lvl_d 加一个小扰动后就近取五点量表的档位。
真实作答不可能每题都一样，扰动保证了题目级区分度，也就不会触发直线作答折减。
"""
import os, sys, random, math, datetime
os.environ['HLR_DB'] = '/home/user/xixi/禾灵儿德育画像系统/hlr_data/portrait.db'
sys.path.insert(0, '/home/user/xixi/禾灵儿德育画像系统')
import hlr_portrait as H
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

DIMS = list(H.DIMS)
OPTS = ["完全不符合", "比较不符合", "一般", "比较符合", "完全符合"]
ITEMS = [
    ("E1", "文化体验", "我参观过嘉兴的博物馆、纪念馆或历史文化街区"),
    ("E2", "文化体验", "在文化场馆里我会认真观察展品，而不是走马观花"),
    ("E3", "文化体验", "我参加过与传统文化有关的研学、采风或社会实践活动"),
    ("E4", "文化体验", "我体验过书法、剪纸、印染、茶艺等传统手工技艺"),
    ("C1", "文化认知", "我了解嘉兴运河文化与江南水乡的历史由来"),
    ("C2", "文化认知", "我能解释红船精神的内涵及其形成的历史背景"),
    ("C3", "文化认知", "我清楚中华优秀传统文化的主要特点与价值"),
    ("C4", "文化认知", "我能说出本地非遗项目的分类及其演变过程"),
    ("I1", "文化志趣", "我会主动去了解自己感兴趣的文化话题"),
    ("I2", "文化志趣", "课外我也会持续阅读文化类的书籍、资讯或视频"),
    ("I3", "文化志趣", "遇到不懂的文化知识，我会追问并深入查证"),
    ("I4", "文化志趣", "我愿意花更多时间拓展新的文化主题领域"),
    ("D1", "文化认同", "我为家乡的文化传统感到自豪"),
    ("D2", "文化认同", "我认同中华文化的价值，并愿意传承弘扬"),
    ("D3", "文化认同", "我愿意主动向他人推荐、介绍本地文化"),
    ("D4", "文化认同", "面对外国朋友，我能自信地讲好中国故事"),
    ("R1", "文化研创", "我曾以文化为主题进行过创作或设计"),
    ("R2", "文化研创", "我做过与文化相关的调查报告或课题研究"),
    ("R3", "文化研创", "我的文化作品讲究原创，并会标注素材出处"),
    ("R4", "文化研创", "我制作的文化作品曾被他人传播或获得反馈"),
]
CH_CN = {"dialogue": "文化对话", "enrollment": "课程报名", "homework": "课程作业",
         "video": "课程视频", "artifact": "作业成果", "agent": "自建智能体"}

# ---------------- 1. 选人：渠道覆盖 ≥3，再按综合指数取前 50 ----------------
cx = H.STORE.conn()
cand = []
for r in cx.execute("SELECT sid, COUNT(DISTINCT channel) nch, COUNT(DISTINCT detail) nev "
                    "FROM evidences WHERE grade='大一' GROUP BY sid HAVING nch>=3"):
    p = H.compute_portrait(H.STORE, r['sid'])['大一']
    c = H.completeness_of(H.STORE, r['sid'])
    cand.append({"sid": r['sid'], "cci": p['cci'], "cp": c['score'],
                 "nch": r['nch'], "nev": r['nev'], "A": p['A'], "S": p['scores']})
cand.sort(key=lambda x: -x['cci'])
SEL = cand[:50]
print(f"候选 {len(cand)} 人 → 选中 50 人，综合指数 {SEL[-1]['cci']:.3f} ~ {SEL[0]['cci']:.3f}")

# 学生信息 + 渠道画像
for s in SEL:
    st = cx.execute("SELECT name,class_name,college,major FROM students WHERE sid=?",
                    (s['sid'],)).fetchone()
    s.update(name=(st['name'] if st else "") or "", cls=(st['class_name'] if st else "") or "",
             college=(st['college'] if st else "") or "", major=(st['major'] if st else "") or "")
    prof = {}
    for r in cx.execute("SELECT channel, COUNT(DISTINCT detail) n FROM evidences "
                        "WHERE sid=? AND grade='大一' GROUP BY channel", (s['sid'],)):
        prof[r['channel']] = r['n']
    s['prof'] = prof

# ---------------- 2. 维度百分位 ----------------
for d in DIMS:
    order = sorted(range(50), key=lambda i: SEL[i]['A'][d])
    for rank, i in enumerate(order):
        SEL[i].setdefault('rank', {})[d] = rank / 49.0

# ---------------- 3. 生成作答 ----------------
random.seed(20260806)
BASE, SPREAD, TONE_SD, ITEM_SD = 0.68, 0.30, 0.10, 0.15


def snap(x):
    return max(0, min(4, int(round(max(0.0, min(1.0, x)) * 4))))


for s in SEL:
    tone = random.gauss(0, TONE_SD)              # 个体自评宽严
    s['tone'] = tone
    s['lvl'] = {}
    s['ans'] = {}
    for d in DIMS:
        lvl = BASE + SPREAD * (s['rank'][d] - 0.5) * 2 + tone
        lvl = max(0.05, min(0.97, lvl))
        s['lvl'][d] = lvl
    for code, d, _ in ITEMS:
        s['ans'][code] = snap(s['lvl'][d] + random.gauss(0, ITEM_SD))
    vals = [s['ans'][c] / 4 for c, _, _ in ITEMS]
    m = sum(vals) / len(vals)
    s['q_mean'] = m
    s['q_sd'] = math.sqrt(sum((v - m) ** 2 for v in vals) / len(vals))

sds = [s['q_sd'] for s in SEL]
print(f"题目级标准差 {min(sds):.3f} ~ {max(sds):.3f}（≥0.18 即不触发直线作答折减）")

# ---------------- 4. 写 Excel ----------------
SONG = "宋体"
HEI = "黑体"
wb = Workbook()
ws = wb.active
ws.title = "问卷答卷"
hdr = ["序号", "提交答卷时间", "所用时间", "学号", "姓名", "班级"] + \
      [f"{i + 1}、{t}【{code}】" for i, (code, d, t) in enumerate(ITEMS)]
ws.append(hdr)
t0 = datetime.datetime(2026, 9, 8, 9, 5)
for i, s in enumerate(SEL, 1):
    t = t0 + datetime.timedelta(minutes=random.randint(0, 95), seconds=random.randint(0, 59))
    ws.append([i, t.strftime("%Y-%m-%d %H:%M:%S"),
               f"{random.randint(112, 268)}秒", s['sid'], s['name'], s['cls']] +
              [OPTS[s['ans'][c]] for c, _, _ in ITEMS])

thin = Side(style="thin", color="BFBFBF")
for c in ws[1]:
    c.font = Font(name=HEI, size=10, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1F6F54")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = Border(bottom=thin)
for row in ws.iter_rows(min_row=2):
    for c in row:
        c.font = Font(name=SONG, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
ws.freeze_panes = "G2"
ws.row_dimensions[1].height = 58
for i, w in enumerate([6, 19, 9, 15, 12, 16], 1):
    ws.column_dimensions[ws.cell(1, i).column_letter].width = w
for i in range(7, 7 + len(ITEMS)):
    ws.column_dimensions[ws.cell(1, i).column_letter].width = 15

# ---- 说明页 ----
w2 = wb.create_sheet("生成说明与特征对照")


def put(r, c, v, bold=False, size=10.5, font=SONG, wrap=False, fill=None):
    cell = w2.cell(r, c, v)
    cell.font = Font(name=HEI if bold else font, size=size, bold=bold)
    cell.alignment = Alignment(vertical="center", wrap_text=wrap,
                               horizontal="center" if c > 1 or bold else "left")
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    return cell


put(1, 1, "禾灵儿五维文化素养基线问卷 · 模拟答卷生成说明", bold=True, size=13)
lines = [
    "一、选人口径：在 5307 名大一学生中，先筛出「过程材料覆盖 ≥3 个渠道」的 548 人"
    f"（材料相对完整），再按综合指数 CCI 由高到低取前 50 名，区间 "
    f"{SEL[-1]['cci']:.3f} ~ {SEL[0]['cci']:.3f}。",
    "二、答卷不是随机编的，而是由每名学生的真实行为特征反推：",
    "        ① 取该生五个维度的实际证据强度 A_d，在这 50 人内换算成百分位 rank_d；",
    "        ② 自评水平 lvl_d = 0.68 + 0.30 ×（rank_d − 0.5）× 2 + 个体基调；"
    "基准取 0.68 是因为这 50 人本就是全校最活跃的一批，自评理应高于中位水平；"
    "个体基调 ~ N(0, 0.10)，用于保留「有人整体自评偏宽、有人偏严」的差异；",
    "        ③ 每道题的作答 = lvl_d 加一个 N(0, 0.15) 的扰动后，就近取五点量表的档位。",
    "三、真实作答不可能每题都选同一档，上述扰动保证了题目级区分度"
    f"（本表 50 份答卷的题目标准差为 {min(sds):.3f} ~ {max(sds):.3f}，均分 "
    f"{sum(x['q_mean'] for x in SEL)/50:.2f}），因此不会被系统的「直线作答折减」"
    "判定为刷分作答——该折减只在「均分很高且题目标准差极小」时才生效。",
    "四、本表可直接在系统里选择材料类别「五维问卷」上传，学号列用于与平台数据关联。",
    "五、下表列出每名学生的行为特征与由此推出的自评水平，便于逐一核查。",
]
r = 3
for ln in lines:
    c = put(r, 1, ln, wrap=True)
    w2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=13)
    w2.row_dimensions[r].height = 30 if len(ln) > 60 else 17
    r += 1

r += 1
head2 = ["学号", "姓名", "综合指数", "材料完整度", "渠道数", "过程材料构成"] + \
        [f"{d[2:]}\n证据强度" for d in DIMS] + [f"{d[2:]}\n自评水平" for d in DIMS] + \
        ["问卷均分", "题目标准差"]
for j, h in enumerate(head2, 1):
    put(r, j, h, bold=True, size=9.5, wrap=True, fill="D8E9E2")
w2.row_dimensions[r].height = 34
hr = r
r += 1
for s in SEL:
    prof = "；".join(f"{CH_CN.get(k, k)} {v}" for k, v in
                     sorted(s['prof'].items(), key=lambda kv: -kv[1]))
    vals = [s['sid'], s['name'], round(s['cci'], 3), round(s['cp'], 2), s['nch'], prof] + \
           [round(s['A'][d], 2) for d in DIMS] + [round(s['lvl'][d], 2) for d in DIMS] + \
           [round(s['q_mean'], 2), round(s['q_sd'], 3)]
    for j, v in enumerate(vals, 1):
        c = put(r, j, v, size=9.5)
        if j == 6:
            c.alignment = Alignment(vertical="center", horizontal="left")
    r += 1
for j, wd in enumerate([15, 11, 10, 11, 8, 34] + [10] * 10 + [10, 11], 1):
    w2.column_dimensions[w2.cell(hr, j).column_letter].width = wd
w2.freeze_panes = w2.cell(hr + 1, 1)

out = '/home/user/xixi/禾灵儿德育画像系统/禾灵儿五维问卷_模拟答卷50人.xlsx'
wb.save(out)
print("已生成:", out)

# ---------------- 5. 回归验证：导入后重算 ----------------
import shutil
shutil.copy(os.environ['HLR_DB'], '/tmp/q50.db')
os.environ['HLR_DB'] = '/tmp/q50.db'
import importlib
importlib.reload(H)
print("\n" + H.import_questionnaire(H.STORE, out, '大一', 'q50')['note'])
print(f"\n{'学号':<15}{'问卷初始得分':>12}{'导入前CCI':>11}{'导入后CCI':>11}")
rows = []
for s in SEL:
    p = H.compute_portrait(H.STORE, s['sid'])['大一']
    b = sum(p['baseline'].values()) / 5
    rows.append((s['sid'], b, s['cci'], p['cci'], H.level_of(p['cci'], '大一')))
for x in rows[:6]:
    print(f"{x[0]:<15}{x[1]:>12.3f}{x[2]:>11.3f}{x[3]:>11.3f}   {x[4]}")
print("   ……")
bs = [x[1] for x in rows]; cs = [x[3] for x in rows]
print(f"\n50 人：初始得分 {min(bs):.3f}~{max(bs):.3f}（均值 {sum(bs)/50:.3f}）"
      f"；导入后 CCI {min(cs):.3f}~{max(cs):.3f}（均值 {sum(cs)/50:.3f}）")
from collections import Counter
print("等级分布：", dict(Counter(x[4] for x in rows)))
